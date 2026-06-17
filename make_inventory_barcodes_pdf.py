from collections import defaultdict
from pathlib import Path

from barcode import Code128
from barcode.writer import ImageWriter
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas


SOURCE = Path(r"D:\Downloads\inventory-20260610-002153.xlsx")
OUTPUT = Path(r"D:\Downloads\inventory-barcodes-by-category.pdf")


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def wrap_text(text, max_chars, max_lines=2):
    words = text.split()
    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if len(candidate) <= max_chars:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word[:max_chars]
        if len(lines) == max_lines:
            break
    if len(lines) < max_lines and current:
        lines.append(current)
    if len(lines) == max_lines and len(" ".join(words)) > len(" ".join(lines)):
        lines[-1] = lines[-1].rstrip(".") + "..."
    return lines


def load_products():
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [clean(v) for v in next(rows)]
    index = {name: i for i, name in enumerate(headers)}

    required = ["Item Code", "Name", "Category"]
    missing = [name for name in required if name not in index]
    if missing:
        raise RuntimeError(f"Missing required columns: {', '.join(missing)}")

    grouped = defaultdict(list)
    for row in rows:
        item_code = clean(row[index["Item Code"]])
        name = clean(row[index["Name"]])
        category = clean(row[index["Category"]]) or "UNCATEGORIZED"
        if not item_code:
            continue
        grouped[category].append({"item_code": item_code, "name": name})
    return dict(sorted(grouped.items(), key=lambda kv: kv[0].casefold()))


def draw_category_header(pdf, category, page_w, page_h, count):
    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(14 * mm, page_h - 16 * mm, category)
    pdf.setFillColor(colors.HexColor("#4b5563"))
    pdf.setFont("Helvetica", 9)
    pdf.drawRightString(page_w - 14 * mm, page_h - 16 * mm, f"{count} products")
    pdf.setStrokeColor(colors.HexColor("#d1d5db"))
    pdf.line(14 * mm, page_h - 20 * mm, page_w - 14 * mm, page_h - 20 * mm)


def make_barcode_image(item_code):
    barcode = Code128(item_code, writer=ImageWriter())
    return barcode.render(
        {
            "module_width": 0.22,
            "module_height": 11,
            "quiet_zone": 2.0,
            "font_size": 0,
            "write_text": False,
            "dpi": 220,
        }
    )


def create_pdf(grouped):
    page_w, page_h = A4
    margin_x = 14 * mm
    gap = 5 * mm
    cols = 2
    card_w = (page_w - (2 * margin_x) - gap) / cols
    card_h = 38 * mm
    top_y = page_h - 26 * mm
    bottom_y = 14 * mm

    pdf = canvas.Canvas(str(OUTPUT), pagesize=A4)
    total = 0

    for category, products in grouped.items():
        y = top_y
        col = 0
        draw_category_header(pdf, category, page_w, page_h, len(products))

        for product in products:
            if y - card_h < bottom_y:
                pdf.showPage()
                y = top_y
                col = 0
                draw_category_header(pdf, category, page_w, page_h, len(products))

            x = margin_x + col * (card_w + gap)
            pdf.setStrokeColor(colors.HexColor("#e5e7eb"))
            pdf.roundRect(x, y - card_h, card_w, card_h, 2 * mm, stroke=1, fill=0)

            img = make_barcode_image(product["item_code"])
            img_w = card_w - 12 * mm
            img_h = 17 * mm
            pdf.drawInlineImage(img, x + 6 * mm, y - 20 * mm, img_w, img_h)

            pdf.setFillColor(colors.HexColor("#111827"))
            pdf.setFont("Helvetica-Bold", 10)
            pdf.drawCentredString(x + card_w / 2, y - 24 * mm, product["item_code"])

            pdf.setFillColor(colors.HexColor("#374151"))
            pdf.setFont("Helvetica", 7.5)
            name_lines = wrap_text(product["name"], 38, 2)
            for line_no, line in enumerate(name_lines):
                pdf.drawCentredString(x + card_w / 2, y - (29 + line_no * 3.5) * mm, line)

            total += 1
            if col == cols - 1:
                col = 0
                y -= card_h + 5 * mm
            else:
                col += 1

        pdf.showPage()

    pdf.save()
    return total


if __name__ == "__main__":
    products_by_category = load_products()
    product_count = create_pdf(products_by_category)
    print(f"Wrote {OUTPUT}")
    print(f"Categories: {len(products_by_category)}")
    print(f"Products: {product_count}")
